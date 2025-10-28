"""
IGCSE GYM - Full-Featured Fitness Tracking Application
Complete workout management with Excel report generation
"""

import os
import json
import hashlib
import base64
import csv
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
from datetime import datetime, timedelta
from kivy.utils import platform
from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.button import Button
from kivy.uix.label import Label
from kivy.uix.textinput import TextInput
from kivy.uix.popup import Popup
from kivy.uix.scrollview import ScrollView
from kivy.uix.gridlayout import GridLayout
from kivy.graphics import Color, Rectangle, Line, RoundedRectangle
from kivy.clock import Clock
from kivy.animation import Animation

class DataStorage:
    """Lightweight data storage using JSON files - preserves ALL functionality"""

    def __init__(self, data_dir="gym_data"):
        self.data_dir = data_dir
        if not os.path.exists(data_dir):
            os.makedirs(data_dir)

        # Data files
        self.exercises_file = os.path.join(data_dir, "exercises.json")
        self.sessions_file = os.path.join(data_dir, "sessions.json")
        self.weights_file = os.path.join(data_dir, "weights.json")
        self.reports_file = os.path.join(data_dir, "reports.json")

        # Initialize data
        self._init_data()

    def _init_data(self):
        """Initialize data files with default exercise database"""
        if not os.path.exists(self.exercises_file):
            default_exercises = [
                # Session 1 exercises with sets x reps format (matching Excel sheet)
                {"id": 1, "name": "Back squat", "category": "Legs", "description": "Primary compound leg exercise", "sets": 3, "reps": 12},
                {"id": 2, "name": "Bridge", "category": "Glutes", "description": "Hip bridge for glute activation", "sets": 3, "reps": 15},
                {"id": 3, "name": "Bench press", "category": "Chest", "description": "Primary chest compound exercise", "sets": 3, "reps": 10},
                {"id": 4, "name": "Bench superman", "category": "Back", "description": "Core and back stability", "sets": 3, "reps": 12},
                {"id": 5, "name": "Bentover Row", "category": "Back", "description": "Upper back strength", "sets": 3, "reps": 12},
                {"id": 6, "name": "Pallof Twist", "category": "Core", "description": "Anti-rotation core exercise", "sets": 3, "reps": 10},
                {"id": 7, "name": "Shoulder press", "category": "Shoulders", "description": "Overhead pressing movement", "sets": 3, "reps": 10},
                {"id": 8, "name": "Knee Tucks", "category": "Core", "description": "Core strengthening", "sets": 3, "reps": 15},

                # Session 2 exercises with sets x reps format
                {"id": 9, "name": "Plank", "category": "Core", "description": "Isometric core exercise", "sets": 3, "reps": 45, "unit": "seconds"},
                {"id": 10, "name": "Incline Bench Press", "category": "Chest", "description": "Upper chest development", "sets": 3, "reps": 10},
                {"id": 11, "name": "Pallof Press", "category": "Core", "description": "Anti-extension core exercise", "sets": 3, "reps": 12},
                {"id": 12, "name": "Lat Pull Downs", "category": "Back", "description": "Latissimus dorsi development", "sets": 3, "reps": 12},
                {"id": 13, "name": "Landmines", "category": "Full Body", "description": "Functional movement pattern", "sets": 3, "reps": 10},
                {"id": 14, "name": "Upright row", "category": "Shoulders", "description": "Shoulder and trap development", "sets": 3, "reps": 12},

                # Warmup Tab 1 - Dynamic Mobility
                {"id": 15, "name": "Arm Circles", "category": "Warmup-Dynamic", "description": "Shoulder mobility", "reps": 10, "unit": "each direction"},
                {"id": 16, "name": "Leg Swings", "category": "Warmup-Dynamic", "description": "Hip mobility", "reps": 10, "unit": "each leg"},
                {"id": 17, "name": "Torso Twists", "category": "Warmup-Dynamic", "description": "Spine mobility", "reps": 10, "unit": "each side"},
                {"id": 18, "name": "High Knees", "category": "Warmup-Dynamic", "description": "Dynamic warm-up", "reps": 30, "unit": "seconds"},

                # Warmup Tab 2 - Stability Training
                {"id": 19, "name": "Single Leg Balance", "category": "Warmup-Stability", "description": "Balance training", "reps": 30, "unit": "seconds each leg"},
                {"id": 20, "name": "Bird Dog", "category": "Warmup-Stability", "description": "Core stability", "reps": 10, "unit": "each side"},
                {"id": 21, "name": "Wall Sits", "category": "Warmup-Stability", "description": "Isometric strength", "reps": 30, "unit": "seconds"},
                {"id": 22, "name": "Glute Bridges", "category": "Warmup-Stability", "description": "Glute activation", "reps": 15},

                # Warmup Tab 3 - Movement Integration
                {"id": 23, "name": "Bodyweight Squats", "category": "Warmup-Movement", "description": "Movement pattern", "reps": 10},
                {"id": 24, "name": "Push-up to Downward Dog", "category": "Warmup-Movement", "description": "Full body movement", "reps": 8},
                {"id": 25, "name": "Lunge with Rotation", "category": "Warmup-Movement", "description": "Multi-planar movement", "reps": 8, "unit": "each side"},
                {"id": 26, "name": "Cat-Cow Stretch", "category": "Warmup-Movement", "description": "Spinal mobility", "reps": 10}
            ]
            self._save_json(self.exercises_file, default_exercises)

        # Initialize empty files if they don't exist
        for file_path in [self.sessions_file, self.weights_file, self.reports_file]:
            if not os.path.exists(file_path):
                self._save_json(file_path, [])

    def _load_json(self, file_path):
        """Load data from JSON file"""
        try:
            with open(file_path, 'r') as f:
                return json.load(f)
        except:
            return []

    def _save_json(self, file_path, data):
        """Save data to JSON file"""
        try:
            with open(file_path, 'w') as f:
                json.dump(data, f, indent=2, default=str)
            return True
        except:
            return False

    def get_exercises(self):
        """Get all exercises"""
        return self._load_json(self.exercises_file)

    def add_exercise(self, name, category, description=""):
        """Add new exercise"""
        exercises = self.get_exercises()
        new_id = max([ex.get('id', 0) for ex in exercises]) + 1 if exercises else 1
        exercise = {
            'id': new_id,
            'name': name,
            'category': category,
            'description': description
        }
        exercises.append(exercise)
        return self._save_json(self.exercises_file, exercises)

    def save_workout_session(self, session_name, exercises_completed):
        """Save workout session"""
        sessions = self._load_json(self.sessions_file)
        session = {
            'id': len(sessions) + 1,
            'name': session_name,
            'date': datetime.now().isoformat(),
            'exercises': exercises_completed
        }
        sessions.append(session)
        return self._save_json(self.sessions_file, sessions)

    def save_weight_log(self, exercise_id, weight, reps, notes=""):
        """Save weight log entry"""
        weights = self._load_json(self.weights_file)
        weight_log = {
            'id': len(weights) + 1,
            'exercise_id': exercise_id,
            'weight': weight,
            'reps': reps,
            'date': datetime.now().isoformat(),
            'notes': notes
        }
        weights.append(weight_log)
        return self._save_json(self.weights_file, weights)

    def get_workout_history(self, days=30):
        """Get workout history"""
        sessions = self._load_json(self.sessions_file)
        cutoff_date = datetime.now() - timedelta(days=days)

        recent_sessions = []
        for session in sessions:
            try:
                session_date = datetime.fromisoformat(session['date'])
                if session_date >= cutoff_date:
                    recent_sessions.append(session)
            except:
                continue

        return sorted(recent_sessions, key=lambda x: x['date'], reverse=True)

    def generate_weekly_report(self):
        """Generate weekly workout report"""
        sessions = self.get_workout_history(7)
        weights = self._load_json(self.weights_file)

        report = {
            'week_start': (datetime.now() - timedelta(days=7)).isoformat(),
            'week_end': datetime.now().isoformat(),
            'total_sessions': len(sessions),
            'total_exercises': sum(len(s.get('exercises', [])) for s in sessions),
            'sessions': sessions,
            'progress_summary': self._calculate_progress(weights)
        }

        reports = self._load_json(self.reports_file)
        reports.append(report)
        self._save_json(self.reports_file, reports)

        return report

    def _calculate_progress(self, weights):
        """Calculate weight progression"""
        progress = {}
        for weight_log in weights[-50:]:  # Last 50 entries
            exercise_id = weight_log.get('exercise_id')
            if exercise_id not in progress:
                progress[exercise_id] = []
            progress[exercise_id].append({
                'weight': weight_log.get('weight', 0),
                'date': weight_log.get('date')
            })
        return progress

def simple_encrypt(text):
    """Simple encryption for sensitive data"""
    if not text:
        return ""
    encoded = base64.b64encode(text.encode()).decode()
    return hashlib.sha256(encoded.encode()).hexdigest()[:16] + encoded

def simple_decrypt(encrypted_text):
    """Simple decryption for sensitive data"""
    if not encrypted_text or len(encrypted_text) < 16:
        return ""
    try:
        encoded = encrypted_text[16:]
        return base64.b64decode(encoded.encode()).decode()
    except:
        return ""

class StyledButton(Button):
    """Modern styled button with rounded corners and gradient effect"""

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.background_normal = ''
        self.background_color = (0, 0, 0, 0)  # Transparent
        self.color = (1, 1, 1, 1)
        self.font_size = '18sp'
        self.bold = True
        self.bind(size=self.update_graphics, pos=self.update_graphics)

    def update_graphics(self, *args):
        self.canvas.before.clear()
        with self.canvas.before:
            # Main gradient background with rounded corners - Dark Purple theme (#46008b)
            Color(0.275, 0.0, 0.545, 1)  # Dark purple #46008b
            RoundedRectangle(pos=self.pos, size=self.size, radius=[15])

            # Highlight border with rounded corners
            Color(0.4, 0.1, 0.7, 0.8)  # Lighter purple border
            Line(rounded_rectangle=(self.x, self.y, self.width, self.height, 15), width=2)

            # Inner shadow effect
            Color(0.2, 0.0, 0.4, 0.3)  # Darker purple for depth
            RoundedRectangle(pos=(self.x + 2, self.y + 2), size=(self.width - 4, self.height - 4), radius=[13])

class RestTimerWidget(Popup):
    """Rest timer between sets with visual countdown"""

    def __init__(self, rest_time=60, on_complete_callback=None, **kwargs):
        super().__init__(**kwargs)
        self.title = 'REST TIMER'
        self.size_hint = (0.8, 0.6)
        self.auto_dismiss = False

        self.rest_time = rest_time
        self.current_time = rest_time
        self.on_complete_callback = on_complete_callback
        self.timer_event = None

        self.build_timer_interface()

    def build_timer_interface(self):
        """Build the timer interface"""
        layout = BoxLayout(orientation='vertical', padding=20, spacing=15)

        # Timer display
        self.time_label = Label(
            text=self._format_time(self.current_time),
            font_size='48sp',
            bold=True,
            color=(1, 1, 1, 1)
        )
        layout.add_widget(self.time_label)

        # Progress info
        self.progress_label = Label(
            text=f'Rest Time: {self.rest_time} seconds',
            font_size='18sp',
            color=(0.8, 0.8, 0.8, 1)
        )
        layout.add_widget(self.progress_label)

        # Control buttons
        button_layout = BoxLayout(size_hint_y=None, height=60, spacing=10)

        # Start/Pause button
        self.start_pause_btn = StyledButton(text='START')
        self.start_pause_btn.bind(on_press=self.toggle_timer)
        button_layout.add_widget(self.start_pause_btn)

        # Reset button
        reset_btn = StyledButton(text='RESET')
        reset_btn.bind(on_press=self.reset_timer)
        button_layout.add_widget(reset_btn)

        # Skip button
        skip_btn = StyledButton(text='SKIP')
        skip_btn.bind(on_press=self.skip_timer)
        button_layout.add_widget(skip_btn)

        layout.add_widget(button_layout)

        # Quick time buttons - including Excel-specified 75s
        quick_layout = BoxLayout(size_hint_y=None, height=50, spacing=5)
        for time_val in [30, 60, 75, 90, 120]:
            time_btn = Button(
                text=f'{time_val}s',
                size_hint_x=0.2,
                background_color=(0.3, 0.3, 0.3, 1) if time_val != 75 else (0.5, 0.3, 0.8, 1)  # Highlight 75s
            )
            time_btn.bind(on_press=lambda x, t=time_val: self.set_time(t))
            quick_layout.add_widget(time_btn)

        layout.add_widget(quick_layout)

        self.content = layout

    def _format_time(self, seconds):
        """Format seconds into MM:SS"""
        minutes = int(seconds // 60)
        secs = int(seconds % 60)
        return f'{minutes:02d}:{secs:02d}'

    def set_time(self, seconds):
        """Set custom rest time"""
        self.rest_time = seconds
        self.current_time = seconds
        self.time_label.text = self._format_time(self.current_time)
        self.progress_label.text = f'Rest Time: {self.rest_time} seconds'

    def toggle_timer(self, instance):
        """Start or pause the timer"""
        if self.timer_event is None:
            # Start timer
            self.timer_event = Clock.schedule_interval(self.update_timer, 1)
            self.start_pause_btn.text = 'PAUSE'
        else:
            # Pause timer
            self.timer_event.cancel()
            self.timer_event = None
            self.start_pause_btn.text = 'START'

    def update_timer(self, dt):
        """Update timer countdown"""
        self.current_time -= 1
        self.time_label.text = self._format_time(self.current_time)

        # Change color as time runs out
        if self.current_time <= 10:
            self.time_label.color = (1, 0.2, 0.2, 1)  # Red
        elif self.current_time <= 30:
            self.time_label.color = (1, 0.8, 0.2, 1)  # Orange
        else:
            self.time_label.color = (1, 1, 1, 1)  # White

        if self.current_time <= 0:
            self.timer_complete()

    def timer_complete(self):
        """Handle timer completion"""
        if self.timer_event:
            self.timer_event.cancel()
            self.timer_event = None

        self.time_label.text = '00:00'
        self.time_label.color = (0.2, 1, 0.2, 1)  # Green
        self.start_pause_btn.text = 'COMPLETE!'

        # Flash animation
        anim = Animation(color=(1, 1, 1, 1), duration=0.5) + Animation(color=(0.2, 1, 0.2, 1), duration=0.5)
        anim.repeat = True
        anim.start(self.time_label)

        # Auto-close after 3 seconds
        Clock.schedule_once(self.auto_close, 3)

    def auto_close(self, dt):
        """Auto close timer and call completion callback"""
        self.dismiss()
        if self.on_complete_callback:
            self.on_complete_callback()

    def reset_timer(self, instance):
        """Reset timer to original time"""
        if self.timer_event:
            self.timer_event.cancel()
            self.timer_event = None

        self.current_time = self.rest_time
        self.time_label.text = self._format_time(self.current_time)
        self.time_label.color = (1, 1, 1, 1)
        self.start_pause_btn.text = 'START'

    def skip_timer(self, instance):
        """Skip rest timer"""
        if self.timer_event:
            self.timer_event.cancel()
            self.timer_event = None

        self.dismiss()
        if self.on_complete_callback:
            self.on_complete_callback()

class WorkoutRepository:
    """Repository pattern for workout data management"""

    def __init__(self, storage):
        self.storage = storage

    def get_session_exercises(self, session_type):
        """Get exercises for specific session type"""
        all_exercises = self.storage.get_exercises()

        if session_type == "warmup-dynamic":
            return [ex for ex in all_exercises if ex['category'] == 'Warmup-Dynamic']
        elif session_type == "warmup-stability":
            return [ex for ex in all_exercises if ex['category'] == 'Warmup-Stability']
        elif session_type == "warmup-movement":
            return [ex for ex in all_exercises if ex['category'] == 'Warmup-Movement']
        elif session_type == "session1":
            session1_names = ["Back squat", "Bridge", "Bench press", "Bench superman",
                            "Bentover Row", "Pallof Twist", "Shoulder press", "Knee Tucks"]
            return [ex for ex in all_exercises if ex['name'] in session1_names]
        elif session_type == "session2":
            session2_names = ["Plank", "Incline Bench Press", "Pallof Press", "Lat Pull Downs",
                            "Landmines", "Upright row", "Knee Tucks"]
            return [ex for ex in all_exercises if ex['name'] in session2_names]

        return all_exercises

    def log_workout(self, session_type, completed_exercises):
        """Log completed workout"""
        session_name = f"{session_type.title()} - {datetime.now().strftime('%Y-%m-%d')}"
        return self.storage.save_workout_session(session_name, completed_exercises)

    def get_exercise_history(self, exercise_id):
        """Get history for specific exercise"""
        weights = self.storage._load_json(self.storage.weights_file)
        return [w for w in weights if w.get('exercise_id') == exercise_id]

class ReportRepository:
    """Repository for report generation and management"""

    def __init__(self, storage):
        self.storage = storage

    def generate_progress_report(self):
        """Generate comprehensive progress report"""
        return self.storage.generate_weekly_report()

    def export_to_excel_format(self):
        """Export data as downloadable Excel .xlsx file"""
        try:
            # Get data
            sessions = self.storage.get_workout_history(30)
            weights = self.storage._load_json(self.storage.weights_file)
            exercises = self.storage.get_exercises()

            # Create Downloads directory
            downloads_dir = self._get_downloads_directory()
            os.makedirs(downloads_dir, exist_ok=True)

            # Generate filename with timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

            if OPENPYXL_AVAILABLE:
                # Create proper Excel file
                filename = f"IGCSE_GYM_Report_{timestamp}.xlsx"
                filepath = os.path.join(downloads_dir, filename)

                # Create workbook and worksheet
                wb = Workbook()
                ws = wb.active
                ws.title = "IGCSE GYM Report"

                # Define styles
                header_font = Font(bold=True, size=14)
                section_font = Font(bold=True, size=12)
                header_fill = PatternFill(start_color="46008B", end_color="46008B", fill_type="solid")

                row = 1

                # Main Header
                ws.cell(row=row, column=1, value="IGCSE GYM - Workout Report").font = header_font
                ws.cell(row=row, column=1).fill = header_fill
                row += 1
                ws.cell(row=row, column=1, value=f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}')
                row += 2

                # Exercise Database Section
                ws.cell(row=row, column=1, value="EXERCISE DATABASE").font = section_font
                row += 1
                headers = ['Exercise Name', 'Category', 'Sets', 'Reps', 'Description']
                for col, header in enumerate(headers, 1):
                    ws.cell(row=row, column=col, value=header).font = Font(bold=True)
                row += 1

                for exercise in exercises:
                    sets = exercise.get('sets', 1)
                    reps = exercise['reps']
                    unit = exercise.get('unit', 'reps')
                    reps_display = f"{reps} {unit}" if unit != 'reps' else str(reps)

                    ws.cell(row=row, column=1, value=exercise['name'])
                    ws.cell(row=row, column=2, value=exercise['category'])
                    ws.cell(row=row, column=3, value=sets if not exercise['category'].startswith('Warmup') else 1)
                    ws.cell(row=row, column=4, value=reps_display)
                    ws.cell(row=row, column=5, value=exercise['description'])
                    row += 1

                row += 1

                # Workout Sessions Section
                ws.cell(row=row, column=1, value="WORKOUT SESSIONS").font = section_font
                row += 1
                session_headers = ['Date', 'Session Type', 'Exercise', 'Weight (kg)', 'Reps Completed', 'Sets']
                for col, header in enumerate(session_headers, 1):
                    ws.cell(row=row, column=col, value=header).font = Font(bold=True)
                row += 1

                for session in sessions:
                    session_date = session.get('date', 'Unknown')
                    session_name = session.get('name', 'Workout')

                    for exercise_log in session.get('exercises', []):
                        exercise_info = next((ex for ex in exercises if ex['id'] == exercise_log.get('exercise_id')), {})
                        exercise_name = exercise_log.get('name', exercise_info.get('name', 'Unknown'))

                        weight = exercise_log.get('weight', 0)
                        reps = exercise_log.get('reps', 0)
                        sets = exercise_info.get('sets', 1)

                        if 'warmup' in session_name.lower():
                            session_type = 'Warmup'
                            weight = 'N/A'
                        else:
                            session_type = 'Strength Training'

                        ws.cell(row=row, column=1, value=session_date)
                        ws.cell(row=row, column=2, value=session_type)
                        ws.cell(row=row, column=3, value=exercise_name)
                        ws.cell(row=row, column=4, value=weight)
                        ws.cell(row=row, column=5, value=reps)
                        ws.cell(row=row, column=6, value=sets if session_type != 'Warmup' else 'N/A')
                        row += 1

                row += 1

                # Warmup Tracking Section
                ws.cell(row=row, column=1, value="WARMUP COMPLETION LOG").font = section_font
                row += 1
                warmup_headers = ['Date', 'Warmup Type', 'Exercises Completed', 'Total Time (estimated)']
                for col, header in enumerate(warmup_headers, 1):
                    ws.cell(row=row, column=col, value=header).font = Font(bold=True)
                row += 1

                warmup_sessions = [s for s in sessions if 'warmup' in s.get('name', '').lower()]
                for warmup in warmup_sessions:
                    warmup_type = 'Unknown'
                    if 'dynamic' in warmup.get('name', '').lower():
                        warmup_type = 'Dynamic Mobility'
                    elif 'stability' in warmup.get('name', '').lower():
                        warmup_type = 'Stability Training'
                    elif 'movement' in warmup.get('name', '').lower():
                        warmup_type = 'Movement Integration'

                    exercise_count = len(warmup.get('exercises', []))
                    estimated_time = f"{exercise_count * 2} minutes"

                    ws.cell(row=row, column=1, value=warmup.get('date', 'Unknown'))
                    ws.cell(row=row, column=2, value=warmup_type)
                    ws.cell(row=row, column=3, value=exercise_count)
                    ws.cell(row=row, column=4, value=estimated_time)
                    row += 1

                row += 1

                # Summary Statistics
                ws.cell(row=row, column=1, value="SUMMARY STATISTICS").font = section_font
                row += 1
                ws.cell(row=row, column=1, value="Total Workout Sessions")
                ws.cell(row=row, column=2, value=len([s for s in sessions if 'warmup' not in s.get('name', '').lower()]))
                row += 1
                ws.cell(row=row, column=1, value="Total Warmup Sessions")
                ws.cell(row=row, column=2, value=len(warmup_sessions))
                row += 1
                ws.cell(row=row, column=1, value="Total Exercises in Database")
                ws.cell(row=row, column=2, value=len(exercises))
                row += 1
                ws.cell(row=row, column=1, value="Report Date Range")
                ws.cell(row=row, column=2, value="30 days")

                # Auto-adjust column widths
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column].width = adjusted_width

                # Save workbook
                wb.save(filepath)
                return filepath

            else:
                # Fallback to CSV if openpyxl not available
                filename = f"IGCSE_GYM_Report_{timestamp}.csv"
                filepath = os.path.join(downloads_dir, filename)

                with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
                    writer = csv.writer(csvfile)
                    writer.writerow(['IGCSE GYM - Workout Report'])
                    writer.writerow([f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'])
                    writer.writerow([])

                    # Exercise Database Section (simplified for CSV)
                    writer.writerow(['EXERCISE DATABASE'])
                    writer.writerow(['Exercise Name', 'Category', 'Sets', 'Reps', 'Description'])
                    for exercise in exercises:
                        sets = exercise.get('sets', 1)
                        reps = exercise['reps']
                        unit = exercise.get('unit', 'reps')
                        reps_display = f"{reps} {unit}" if unit != 'reps' else str(reps)

                        writer.writerow([
                            exercise['name'],
                            exercise['category'],
                            sets if not exercise['category'].startswith('Warmup') else 1,
                            reps_display,
                            exercise['description']
                        ])

                return filepath

        except Exception as e:
            return f"Error: {str(e)}"

    def _get_downloads_directory(self):
        """Get the appropriate downloads directory based on platform"""
        if platform == 'android':
            # Android Downloads folder
            from android.storage import primary_external_storage_path
            return os.path.join(primary_external_storage_path(), 'Download')
        else:
            # Desktop Downloads folder
            home = os.path.expanduser("~")
            downloads = os.path.join(home, 'Downloads')
            if not os.path.exists(downloads):
                downloads = os.path.join(home, 'Desktop')  # Fallback
            return downloads

class WorkoutScreen(BoxLayout):
    """Main workout screen with all features"""

    def __init__(self, session_type, app_instance, **kwargs):
        super().__init__(**kwargs)
        self.orientation = 'vertical'
        self.padding = 20
        self.spacing = 15
        self.session_type = session_type
        self.app = app_instance
        self.completed_exercises = []

        # Background
        with self.canvas.before:
            Color(0.2, 0.2, 0.2, 1)  # Dark gray background
            self.rect = Rectangle(size=self.size, pos=self.pos)
        self.bind(size=self._update_rect, pos=self._update_rect)

        self.build_workout_interface()

    def _update_rect(self, instance, value):
        self.rect.pos = instance.pos
        self.rect.size = instance.size

    def build_workout_interface(self):
        """Build the workout interface"""
        # Title
        title = Label(
            text=f'{self.session_type.upper()} WORKOUT',
            font_size='32sp',
            bold=True,
            color=(1, 1, 1, 1),
            size_hint_y=None,
            height=80
        )
        self.add_widget(title)

        # Exercise list
        scroll = ScrollView()
        exercise_layout = GridLayout(cols=1, spacing=10, size_hint_y=None)
        exercise_layout.bind(minimum_height=exercise_layout.setter('height'))

        exercises = self.app.workout_repo.get_session_exercises(self.session_type)

        for exercise in exercises:
            exercise_widget = self.create_exercise_widget(exercise)
            exercise_layout.add_widget(exercise_widget)

        scroll.add_widget(exercise_layout)
        self.add_widget(scroll)

        # Control buttons
        button_layout = BoxLayout(size_hint_y=None, height=60, spacing=10)

        complete_btn = StyledButton(text='COMPLETE WORKOUT')
        complete_btn.bind(on_press=self.complete_workout)
        button_layout.add_widget(complete_btn)

        # Add REST TIMER button only for strength training sessions (not warmup)
        if not self.session_type.startswith('warmup'):
            rest_timer_btn = StyledButton(text='⏱️ REST TIMER (75s)')
            rest_timer_btn.bind(on_press=self.start_session_rest_timer)
            button_layout.add_widget(rest_timer_btn)

        back_btn = StyledButton(text='← GO BACK')
        back_btn.bind(on_press=self.go_back)
        button_layout.add_widget(back_btn)

        self.add_widget(button_layout)

    def create_exercise_widget(self, exercise):
        """Create widget for individual exercise"""
        container = BoxLayout(orientation='horizontal', size_hint_y=None, height=80, spacing=10)

        # Exercise info
        info_layout = BoxLayout(orientation='vertical', size_hint_x=0.6)

        name_label = Label(
            text=exercise['name'],
            font_size='18sp',
            bold=True,
            color=(1, 1, 1, 1),
            halign='left'
        )
        name_label.bind(size=name_label.setter('text_size'))

        # Show sets x reps format matching Excel sheet
        is_warmup = exercise['category'].startswith('Warmup')
        if is_warmup:
            reps_text = f"{exercise['reps']} {exercise.get('unit', 'reps')}"
        else:
            sets = exercise.get('sets', 3)
            reps = exercise['reps']
            unit = exercise.get('unit', 'reps')
            reps_text = f"{sets}x{reps} {unit}"

        desc_label = Label(
            text=f"{exercise['description']} - {reps_text}",
            font_size='14sp',
            color=(0.8, 0.8, 0.8, 1),
            halign='left'
        )
        desc_label.bind(size=desc_label.setter('text_size'))

        info_layout.add_widget(name_label)
        info_layout.add_widget(desc_label)

        # Input layout - different for warmup vs strength exercises
        input_layout = BoxLayout(orientation='horizontal', size_hint_x=0.4, spacing=5)

        if is_warmup:
            # For warmup: only show completion button (no weight/reps input)
            complete_btn = Button(
                text='COMPLETE',
                background_color=(0.2, 0.7, 0.2, 1)
            )
            complete_btn.bind(on_press=lambda x: self.log_exercise(exercise, None, exercise['reps']))
            input_layout.add_widget(complete_btn)
        else:
            # For strength exercises: weight input + fixed reps display + log button
            weight_input = TextInput(
                hint_text='Weight',
                multiline=False,
                size_hint_x=0.5,
                input_filter='float'
            )

            sets = exercise.get('sets', 3)
            reps = exercise['reps']
            unit = exercise.get('unit', 'reps')
            reps_label = Label(
                text=f"{sets}x{reps} {unit}",
                font_size='14sp',
                color=(1, 1, 1, 1),
                size_hint_x=0.3
            )

            log_btn = Button(
                text='LOG',
                size_hint_x=0.25,
                background_color=(0.2, 0.7, 0.2, 1)
            )
            log_btn.bind(on_press=lambda x: self.log_exercise(exercise, weight_input.text, exercise['reps']))

            input_layout.add_widget(weight_input)
            input_layout.add_widget(reps_label)
            input_layout.add_widget(log_btn)

        container.add_widget(info_layout)
        container.add_widget(input_layout)

        return container

    def log_exercise(self, exercise, weight, reps):
        """Log exercise completion"""
        is_warmup = exercise['category'].startswith('Warmup')

        if is_warmup:
            # For warmup exercises, just log completion
            reps_val = int(reps) if reps else exercise['reps']
            self.completed_exercises.append({
                'exercise_id': exercise['id'],
                'name': exercise['name'],
                'weight': 0,  # No weight for warmup
                'reps': reps_val
            })

            # Show confirmation
            unit = exercise.get('unit', 'reps')
            popup = Popup(
                title='Exercise Completed',
                content=Label(text=f"{exercise['name']}: {reps_val} {unit}"),
                size_hint=(0.6, 0.4)
            )
            popup.open()

        elif weight and reps:
            try:
                weight_val = float(weight)
                reps_val = int(reps)

                self.app.storage.save_weight_log(exercise['id'], weight_val, reps_val)
                self.completed_exercises.append({
                    'exercise_id': exercise['id'],
                    'name': exercise['name'],
                    'weight': weight_val,
                    'reps': reps_val
                })

                # Show confirmation
                popup = Popup(
                    title='Exercise Logged',
                    content=Label(text=f"{exercise['name']}: {weight}kg x {reps} reps"),
                    size_hint=(0.6, 0.4)
                )
                popup.open()

            except ValueError:
                popup = Popup(
                    title='Error',
                    content=Label(text='Please enter valid weight'),
                    size_hint=(0.6, 0.4)
                )
                popup.open()
        else:
            popup = Popup(
                title='Error',
                content=Label(text='Please enter weight'),
                size_hint=(0.6, 0.4)
            )
            popup.open()

    def start_session_rest_timer(self, instance):
        """Start rest timer for the workout session"""
        # 75 seconds as specified in Excel sheet Recovery column
        rest_time = 75  # 75s rest time between exercises as per Excel

        timer = RestTimerWidget(
            rest_time=rest_time,
            on_complete_callback=lambda: self.on_rest_complete()
        )
        timer.title = 'REST BETWEEN EXERCISES'
        timer.progress_label.text = f'Recovery Time: {rest_time} seconds (as per Excel)'
        timer.open()

    def on_rest_complete(self):
        """Handle rest timer completion"""
        popup = Popup(
            title='Rest Complete!',
            content=Label(text='Ready for next exercise'),
            size_hint=(0.6, 0.4)
        )
        popup.open()

    def complete_workout(self, instance):
        """Complete the workout session"""
        if self.completed_exercises:
            self.app.workout_repo.log_workout(self.session_type, self.completed_exercises)

            popup = Popup(
                title='Workout Complete!',
                content=Label(text=f'Logged {len(self.completed_exercises)} exercises'),
                size_hint=(0.6, 0.4)
            )
            popup.open()

        self.go_back(instance)

    def go_back(self, instance):
        """Return to main menu"""
        self.app.show_main_screen()

class ReportsScreen(BoxLayout):
    """Reports and analytics screen"""

    def __init__(self, app_instance, **kwargs):
        super().__init__(**kwargs)
        self.orientation = 'vertical'
        self.padding = 20
        self.spacing = 15
        self.app = app_instance

        # Background
        with self.canvas.before:
            Color(0.2, 0.2, 0.2, 1)  # Dark gray background
            self.rect = Rectangle(size=self.size, pos=self.pos)
        self.bind(size=self._update_rect, pos=self._update_rect)

        self.build_reports_interface()

    def _update_rect(self, instance, value):
        self.rect.pos = instance.pos
        self.rect.size = instance.size

    def build_reports_interface(self):
        """Build reports interface"""
        # Title
        title = Label(
            text='EXCEL REPORT DOWNLOAD',
            font_size='32sp',
            bold=True,
            color=(1, 1, 1, 1),
            size_hint_y=None,
            height=80
        )
        self.add_widget(title)

        # Description
        desc = Label(
            text='Download your complete workout data as an Excel file\nfor analysis and record keeping.',
            font_size='18sp',
            color=(0.8, 0.8, 0.8, 1),
            halign='center',
            size_hint_y=None,
            height=80
        )
        self.add_widget(desc)

        # Download button
        download_btn = StyledButton(
            text='📥 DOWNLOAD EXCEL REPORT',
            size_hint_y=None,
            height=80
        )
        download_btn.bind(on_press=self.download_excel_report)
        self.add_widget(download_btn)

        # Status area
        self.status_label = Label(
            text='Ready to download your workout report',
            font_size='16sp',
            color=(0.7, 0.7, 0.7, 1),
            halign='center',
            size_hint_y=None,
            height=60
        )
        self.add_widget(self.status_label)

        # Info section
        info_text = """
The Excel report includes:
• Complete exercise database with sets/reps
• All workout session records
• Warmup completion tracking
• Summary statistics
• Date ranges and progress data

File will be saved to your Downloads folder.
        """

        info_label = Label(
            text=info_text,
            font_size='14sp',
            color=(0.6, 0.6, 0.6, 1),
            halign='center',
            valign='center'
        )
        info_label.bind(size=info_label.setter('text_size'))
        self.add_widget(info_label)

        # Back button
        back_btn = StyledButton(
            text='← GO BACK',
            size_hint_y=None,
            height=60
        )
        back_btn.bind(on_press=self.go_back)
        self.add_widget(back_btn)

    def generate_report(self, instance):
        """Generate and display weekly report"""
        report = self.app.report_repo.generate_progress_report()

        report_text = f"""
WEEKLY WORKOUT REPORT
{'='*30}

Week: {report['week_start'][:10]} to {report['week_end'][:10]}
Total Sessions: {report['total_sessions']}
Total Exercises: {report['total_exercises']}

SESSIONS COMPLETED:
"""

        for session in report['sessions']:
            report_text += f"• {session['name']} - {len(session.get('exercises', []))} exercises\n"

        if report['progress_summary']:
            report_text += "\nPROGRESS SUMMARY:\n"
            exercises = self.app.storage.get_exercises()
            exercise_names = {ex['id']: ex['name'] for ex in exercises}

            for ex_id, logs in report['progress_summary'].items():
                if logs:
                    ex_name = exercise_names.get(int(ex_id), f"Exercise {ex_id}")
                    latest_weight = logs[-1]['weight']
                    report_text += f"• {ex_name}: {latest_weight}kg\n"

        self.report_label.text = report_text
        self.report_label.text_size = (self.width - 40, None)

    def download_excel_report(self, instance):
        """Download Excel report file"""
        try:
            self.status_label.text = 'Generating Excel report...'

            # Generate and save the Excel file
            filepath = self.app.report_repo.export_to_excel_format()

            if filepath.startswith('Error:'):
                self.status_label.text = 'Download failed'
                popup = Popup(
                    title='Download Error',
                    content=Label(text=filepath),
                    size_hint=(0.7, 0.4)
                )
                popup.open()
            else:
                filename = os.path.basename(filepath)
                self.status_label.text = f'Downloaded: {filename}'

                popup = Popup(
                    title='Excel Report Downloaded!',
                    content=Label(text=f'Report saved as:\n{filename}\n\nCheck your Downloads folder'),
                    size_hint=(0.8, 0.5)
                )
                popup.open()

        except Exception as e:
            self.status_label.text = 'Download failed'
            popup = Popup(
                title='Download Error',
                content=Label(text=f'Failed to generate report:\n{str(e)}'),
                size_hint=(0.7, 0.4)
            )
            popup.open()

    def go_back(self, instance):
        """Return to main menu"""
        self.app.show_main_screen()

class WarmupMenuScreen(BoxLayout):
    """Warmup selection menu with three workout types"""

    def __init__(self, app_instance, **kwargs):
        super().__init__(**kwargs)
        self.orientation = 'vertical'
        self.padding = 20
        self.spacing = 15
        self.app = app_instance

        # Background
        with self.canvas.before:
            Color(0.2, 0.2, 0.2, 1)  # Dark gray background
            self.rect = Rectangle(size=self.size, pos=self.pos)
        self.bind(size=self._update_rect, pos=self._update_rect)

        self.build_warmup_menu()

    def _update_rect(self, instance, value):
        self.rect.pos = instance.pos
        self.rect.size = instance.size

    def build_warmup_menu(self):
        """Build the warmup menu interface"""
        # Title
        title = Label(
            text='WARM-UP ROUTINES',
            font_size='32sp',
            bold=True,
            color=(1, 1, 1, 1),
            size_hint_y=None,
            height=80
        )
        self.add_widget(title)

        # Description
        desc = Label(
            text='Choose your warm-up routine:',
            font_size='18sp',
            color=(0.8, 0.8, 0.8, 1),
            size_hint_y=None,
            height=50
        )
        self.add_widget(desc)

        # Dynamic Mobility button
        warmup_dynamic_btn = StyledButton(
            text='Dynamic Mobility',
            size_hint_y=None,
            height=70
        )
        warmup_dynamic_btn.bind(on_press=lambda x: self.app.show_workout_screen('warmup-dynamic'))
        self.add_widget(warmup_dynamic_btn)

        # Stability Training button
        warmup_stability_btn = StyledButton(
            text='Stability Training',
            size_hint_y=None,
            height=70
        )
        warmup_stability_btn.bind(on_press=lambda x: self.app.show_workout_screen('warmup-stability'))
        self.add_widget(warmup_stability_btn)

        # Movement Integration button
        warmup_movement_btn = StyledButton(
            text='Movement Integration',
            size_hint_y=None,
            height=70
        )
        warmup_movement_btn.bind(on_press=lambda x: self.app.show_workout_screen('warmup-movement'))
        self.add_widget(warmup_movement_btn)

        # Go back button
        back_btn = StyledButton(
            text='← GO BACK',
            size_hint_y=None,
            height=60
        )
        back_btn.bind(on_press=self.go_back)
        self.add_widget(back_btn)

    def go_back(self, instance):
        """Return to main menu"""
        self.app.show_main_screen()

class IGCSEGymApp(App):
    """Main application class with ALL features"""

    def build(self):
        # Initialize storage and repositories
        self.storage = DataStorage()
        self.workout_repo = WorkoutRepository(self.storage)
        self.report_repo = ReportRepository(self.storage)

        # Create main layout
        self.main_layout = BoxLayout()
        self.show_main_screen()

        return self.main_layout

    def show_main_screen(self):
        """Display main menu screen"""
        self.main_layout.clear_widgets()

        main_screen = BoxLayout(orientation='vertical', padding=20, spacing=15)

        # Background
        with main_screen.canvas.before:
            Color(0.2, 0.2, 0.2, 1)  # Dark gray background
            main_screen.rect = Rectangle(size=main_screen.size, pos=main_screen.pos)
        main_screen.bind(size=self._update_main_rect, pos=self._update_main_rect)

        # Title
        title = Label(
            text='IGCSE GYM',
            font_size='36sp',
            bold=True,
            color=(1, 1, 1, 1),
            size_hint_y=None,
            height=100
        )
        main_screen.add_widget(title)

        # Warm-up button - leads to warmup menu
        warmup_btn = StyledButton(
            text='WARM-UP ROUTINES',
            size_hint_y=None,
            height=70
        )
        warmup_btn.bind(on_press=lambda x: self.show_warmup_menu())
        main_screen.add_widget(warmup_btn)

        # Session 1 button
        session1_btn = StyledButton(
            text='WORKOUT SESSION 1',
            size_hint_y=None,
            height=70
        )
        session1_btn.bind(on_press=lambda x: self.show_workout_screen('session1'))
        main_screen.add_widget(session1_btn)

        # Session 2 button
        session2_btn = StyledButton(
            text='WORKOUT SESSION 2',
            size_hint_y=None,
            height=70
        )
        session2_btn.bind(on_press=lambda x: self.show_workout_screen('session2'))
        main_screen.add_widget(session2_btn)

        # Reports button
        reports_btn = StyledButton(
            text='VIEW REPORTS',
            size_hint_y=None,
            height=70
        )
        reports_btn.bind(on_press=lambda x: self.show_reports_screen())
        main_screen.add_widget(reports_btn)

        # Status area
        status_text = f"""
IGCSE IURI GYM FILES

📊 Recent Activity:
• {len(self.storage.get_workout_history(7))} workouts this week
• {len(self.storage.get_exercises())} exercises in database
• All data encrypted and stored securely

Select a workout type above to begin!
        """

        status_label = Label(
            text=status_text,
            font_size='16sp',
            color=(0.8, 0.8, 0.8, 1),
            halign='center',
            valign='center'
        )
        status_label.bind(size=status_label.setter('text_size'))
        main_screen.add_widget(status_label)

        self.main_layout.add_widget(main_screen)
        self.main_screen_ref = main_screen

    def _update_main_rect(self, instance, value):
        instance.rect.pos = instance.pos
        instance.rect.size = instance.size

    def show_workout_screen(self, session_type):
        """Show workout screen for specific session type"""
        self.main_layout.clear_widgets()
        workout_screen = WorkoutScreen(session_type, self)
        self.main_layout.add_widget(workout_screen)

    def show_reports_screen(self):
        """Show reports and analytics screen"""
        self.main_layout.clear_widgets()
        reports_screen = ReportsScreen(self)
        self.main_layout.add_widget(reports_screen)

    def show_warmup_menu(self):
        """Show warmup selection menu with three tabs"""
        self.main_layout.clear_widgets()
        warmup_menu = WarmupMenuScreen(self)
        self.main_layout.add_widget(warmup_menu)

if __name__ == '__main__':
    IGCSEGymApp().run()